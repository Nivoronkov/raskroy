"""
Тест разбора спецификации (защита от регрессий).

Зачем: когда дорабатываешь конвертер под новые спецификации (исполнения,
позиции, новые профили), этот тест проверяет, что разбор уже известной
спецификации "Каркас" НЕ сломался.

Запуск (из папки C:\\raskroy, при активированной venv):
    pytest

Если после твоей правки тест "покраснел" — значит правка сломала разбор,
который раньше работал. Смотри, что именно разошлось с эталоном.
"""
import os
import sys
import pandas as pd
import pytest

# подключаем конвертер из родительской папки
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

from sp_to_cutlist import main as convert_sp

ETALON = os.path.join(HERE, "data", "karkas_etalon.xls")


@pytest.fixture(scope="module")
def sheets(tmp_path_factory):
    """Конвертирует эталон один раз и отдаёт листы как DataFrame."""
    out = tmp_path_factory.mktemp("out") / "karkas_вход.xlsx"
    convert_sp(ETALON, str(out))
    return {
        "Детали": pd.read_excel(out, sheet_name="Детали", dtype=str, header=2).fillna(""),
        "Материалы": pd.read_excel(out, sheet_name="Материалы", dtype=str).fillna(""),
        "Проверка": pd.read_excel(out, sheet_name="Проверка", dtype=str).fillna(""),
        "Отсечено": pd.read_excel(out, sheet_name="Отсечено", dtype=str).fillna(""),
    }


def test_число_деталей(sheets):
    # 25 строк линейного проката
    assert len(sheets["Детали"]) == 25


def test_число_материалов(sheets):
    # уникальные материалы (типоразмер + марка из столбца G "Примечание")
    assert len(sheets["Материалы"]) == 6


def test_отсечены_листы(sheets):
    # 20 строк не для линейного раскроя (листы, пластины, крепёж, минвата)
    assert len(sheets["Отсечено"]) == 20


def test_длина_извлечена_у_всех(sheets):
    # ни у одной детали проката длина не должна быть пустой
    дет = sheets["Детали"]
    пустые = дет[дет["Длина, мм"] == ""]
    assert len(пустые) == 0, f"Длина не извлечена у строк: {пустые['Обозначение'].tolist()}"


def test_конкретная_длина_трубы(sheets):
    # поз. ...065-2750 -> труба 50х50х4, длина 2750, количество 8
    дет = sheets["Детали"]
    строка = дет[дет["Обозначение"].str.contains("065-2750", na=False)]
    assert len(строка) == 1
    assert строка.iloc[0]["Длина, мм"] == "2750"
    assert строка.iloc[0]["Количество"] == "8"


def test_труба_марка_из_примечания(sheets):
    # марка берётся из столбца G (Примечание): труба 100х60х5 с С345-5 -> С345,
    # отдельным материалом, не сливается с трубами С255
    мат = sheets["Материалы"]
    assert any("ТР-П-100х60х5-С345" in c for c in мат["Код материала"])
    assert any("С255" in c for c in мат["Код материала"])


def test_марка_из_G_приоритет(sheets):
    # все марки в Каркасе берутся из G (С255/С345), а не из обозначения материала
    мат = sheets["Материалы"]
    assert all(("С255" in c or "С345" in c) for c in мат["Код материала"])


# ============ Эталон "Основание" (доработка, протягивание, конфликты) ============

ETALON_OSN = os.path.join(HERE, "data", "osnovanie_etalon.xls")


@pytest.fixture(scope="module")
def osn(tmp_path_factory):
    out = tmp_path_factory.mktemp("out2") / "osn_вход.xlsx"
    convert_sp(ETALON_OSN, str(out))
    return {
        "Детали": pd.read_excel(out, sheet_name="Детали", dtype=str, header=2).fillna(""),
        "Проверка": pd.read_excel(out, sheet_name="Проверка", dtype=str).fillna(""),
    }


def test_осн_протягивание_обозначения(osn):
    # сокращённая -218_д должна развернуться в полное МРКЕ.746611.014-218_д
    дет = osn["Детали"]
    assert any("746611.014-218_д" in o for o in дет["Обозначение"])


def test_осн_доработка_извлечена(osn):
    # у -218_д1 длина 218 (раньше "не извлечена")
    дет = osn["Детали"]
    строка = дет[дет["Обозначение"].str.contains("218_д1", na=False)]
    assert len(строка) == 1
    assert строка.iloc[0]["Длина, мм"] == "218"
    assert строка.iloc[0]["Доработка"] == "_д1"


def test_осн_марка_из_примечания(osn):
    # размер берётся из наименования, марка — из G; столбец "Обозначение
    # материала" (после G) игнорируется, т.к. часто содержит неверные данные
    дет = osn["Детали"]
    # деталь швеллер 16П должна иметь код именно с 16П (а не с ошибочным 12У из столбца 15)
    assert any("ШВ-16П" in str(c) for c in дет["Код материала"])
    assert not any("12У" in str(c) for c in дет["Код материала"])


# ============ Эталон с исполнениями ============

ETALON_ISP = os.path.join(HERE, "data", "ispolneniya_etalon.xls")


@pytest.fixture(scope="module")
def isp(tmp_path_factory):
    out = tmp_path_factory.mktemp("out3") / "isp_вход.xlsx"
    convert_sp(ETALON_ISP, str(out))
    return {
        "Детали": pd.read_excel(out, sheet_name="Детали", dtype=str, header=2).fillna(""),
        "Материалы": pd.read_excel(out, sheet_name="Материалы", dtype=str).fillna(""),
    }


def test_исп_развёрнуты(isp):
    # деталь -3500 используется в 2 исполнениях -> минимум 2 строки с этим обозначением
    дет = isp["Детали"]
    s3500 = дет[дет["Обозначение"].str.contains("061-3500", na=False)]
    assert len(s3500) >= 2
    исполнения = set(s3500["Исполнение"])
    assert len(исполнения) >= 2  # разные исполнения


def test_исп_номер_в_примечании(isp):
    # КРАТКОЕ примечание (для схемы раскроя): номер исполнения без слова 'исп.'
    дет = isp["Детали"]
    # хотя бы у одной детали краткое примечание содержит номер исполнения
    assert any(str(p).strip() and "исп." not in str(p) for p in дет["Примечание"])
    # ПОЛНОЕ описание (для сводок) сохраняет 'исп.N'
    if "Описание" in дет.columns:
        assert any("исп." in str(p) for p in дет["Описание"])


def test_исп_марка_из_примечания(isp):
    # в этом файле нет сортамента — марка С255/С345 берётся из примечания
    мат = isp["Материалы"]
    марки = set(мат["Марка"])
    assert "С255" in марки or "С345" in марки


# ====== Формат «вся суть в наименовании» (Основание, ещё одна сп) ======

def _convert(path):
    import tempfile
    from openpyxl import load_workbook
    from sp_to_cutlist import main as convert
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    convert(path, tmp.name)
    return load_workbook(tmp.name, data_only=True)


def test_наимформат_длина_из_наименования():
    """Длина берётся из 'L = NNN мм' в наименовании, когда обозначение пустое."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie2_nameformat_etalon.xls"))
    ws = wb["Детали"]
    # ищем швеллер 16П длиной 140 (есть в файле) и проверяем, что длина извлечена
    found = False
    for r in range(4, ws.max_row + 1):
        code = ws.cell(r, 3).value
        length = ws.cell(r, 4).value
        if code == "ШВ-16П-С355" and length == 140:
            found = True
            break
    assert found, "длина из 'L = 140 мм' не извлечена"


def test_наимформат_марка_слипшаяся_с_гостом():
    """Марка С355 распознаётся, хотя слиплась с номером ГОСТа (8240-97С355)."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie2_nameformat_etalon.xls"))
    ws = wb["Материалы"]
    codes = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    assert "ШВ-24П-С355" in codes
    assert "УГ-63х63х5-С255" in codes


def test_наимформат_составная_деталь_помечена():
    """Швеллер 24П длиной 13500 мм (> хлыста) помечается как составная."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie2_nameformat_etalon.xls"))
    ws = wb["Проверка"]
    flagged = False
    for r in range(1, ws.max_row + 1):
        row = " ".join(str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1))
        if "СОСТАВНАЯ ДЕТАЛЬ" in row and "13500" in row:
            flagged = True
            break
    assert flagged, "составная деталь 13500 мм не помечена"


def test_наимформат_круглая_труба_отсечена():
    """Труба 25х2,8 (круглая ВГП) не попадает в раскрой."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie2_nameformat_etalon.xls"))
    ws = wb["Детали"]
    for r in range(4, ws.max_row + 1):
        code = ws.cell(r, 3).value
        assert code != "ТР-П-25х2-С255" and not (code or "").startswith("ТР-П-25х2")


# ====== Эталон с данными после G и круглыми трубами (Основание 059) ======

def test_059_марка_из_G_не_из_столбца15():
    """Марка берётся из G; неверные данные столбца 'Обозначение материала' игнорируются."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie3_kolonki_etalon.xls"))
    ws = wb["Материалы"]
    codes = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    # швеллер 16П с маркой С345 из G (в столбце 15 ошибочно стоял 12У / Сталь 10)
    assert "ШВ-16П-С345" in codes
    # труба 40х40х2 — С255 из G (в столбце 15 стояла "Сталь 10")
    assert "ТР-П-40х40х2-С255" in codes


def test_059_круглые_трубы_раскраиваются():
    """Круглые/ВГП трубы (Ду, ГОСТ 3262) попадают в раскрой как ТР-К."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie3_kolonki_etalon.xls"))
    ws = wb["Материалы"]
    codes = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    assert any(str(c).startswith("ТР-К-80х4") for c in codes)
    assert any(str(c).startswith("ТР-К-40х3") for c in codes)


def test_059_позиции_заголовок_профиля():
    """Позиции 40,41 (профиль в строке-заголовке, наимен. только 'L=184') попадают."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie3_kolonki_etalon.xls"))
    ws = wb["Детали"]
    found = False
    for r in range(4, ws.max_row + 1):
        code = str(ws.cell(r, 3).value or "")
        length = ws.cell(r, 4).value
        if code.startswith("ТР-К") and length == 184:
            found = True
            break
    assert found, "позиции 40/41 (L=184) не попали в раскрой"


def test_059_круглая_труба_без_марки_ст3():
    """Круглой трубе без марки в СП присваивается Ст3 по умолчанию."""
    wb = _convert(os.path.join(HERE, "data", "osnovanie3_kolonki_etalon.xls"))
    ws = wb["Материалы"]
    codes = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    assert any(str(c).startswith("ТР-К") and c.endswith("Ст3") for c in codes)
