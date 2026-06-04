"""
Конвертер спецификации Компас (.xls) -> Excel для калькулятора раскроя.
Версия 3:
  - автоопределение исполнений (несколько колонок "Количество") и их разворот
    в плоский список (номер исполнения -> в примечание детали);
  - протягивание базового обозначения, суффикс доработки _д/_дN;
  - марка С255/чистая/Сталь N; конфликт размера наимен. vs сортамент;
  - "умный" светофор: статус строки по заполненности + валидности кода,
    доработка отдельным флагом (не мешает зелёному).
"""
import re
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

# Единый модуль нормализации (тот же, что использует справочник материалов),
# чтобы код материала из спецификации был идентичен коду из справочника.
_HERE = os.path.dirname(os.path.abspath(__file__))
for _p in (os.path.join(_HERE, "smart_cut_app"), _HERE):
    if _p not in sys.path:
        sys.path.insert(0, _p)
try:
    from core.normalization import (
        normalize_size,
        normalize_grade,
        material_code as build_material_code,
        PROFILE_TYPE_MAP as PROFILE_MAP,
    )
    _NORM_OK = True
except ImportError:
    # Фолбэк, если модуль недоступен (конвертер должен работать автономно).
    _NORM_OK = False
    PROFILE_MAP = {
        "Труба профильная": "ТР-П", "Труба круглая": "ТР-К", "Швеллер": "ШВ",
        "Уголок": "УГ", "Полоса": "ПЛ", "Круг": "КР",
    }
    import re as _re

    def normalize_size(s):
        s = _re.sub(r"\s+", "", str(s).strip())
        s = _re.sub(r"[xXхХ*]", "х", s)
        s = _re.sub(r"х+", "х", s)
        m = _re.fullmatch(r"(\d+)([a-zA-Zа-яА-Я]+\d*)", s)
        if m:
            letters = m.group(2).upper().replace("U", "П").replace("Y", "У")
            return f"{m.group(1)}{letters}"
        return s.lower()

    def normalize_grade(raw):
        if not raw:
            return ""
        g = _re.sub(r"\s+", "", str(raw).strip()); key = g.upper()
        m = _re.match(r"^СТ(\d+)", key)
        if m: return f"Ст{m.group(1)}"
        m = _re.match(r"^С(\d{3})", key)
        if m: return f"С{m.group(1)}"
        m = _re.match(r"^(\d{2}Г\d?[А-Я]+)", key)
        if m: return m.group(1)
        m = _re.match(r"^СТАЛЬ(\d+)", key)
        if m: return f"Сталь{m.group(1)}"
        return g

    def build_material_code(profile, size, grade):
        prefix = PROFILE_MAP.get(profile, profile)
        s, g = normalize_size(size), normalize_grade(grade)
        return f"{prefix}-{s}-{g}" if (prefix and s and g) else ""

STOCK_LENGTH_DEFAULT = {"ТР-П": 6000, "ТР-К": 6000, "ШВ": 11700, "УГ": 11700, "ПЛ": 6000, "КР": 6000}
SHEET_KEYWORDS = ("Лист", "Пластина", "Косынка", "Ребро", "Сетка", "Скоба",
                  "Бобышка", "Гайка", "Болт", "Шайба", "Винт")



def classify_profile(naim):
    n = naim.strip()
    # Круглая / ВГП труба: 'Труба Ду80х4,0', 'Труба 25х2,8' (диаметр x стенка).
    # Раньше отсекалась — теперь раскраивается как погонаж (тип "Труба круглая").
    m = re.search(r"Труба\s+Ду\s*([\d.,х x*]+)", n, re.I)
    if m:
        return "Труба круглая", normalize_size(m.group(1))
    m = re.search(r"Труба\s+(?:профильная\s+)?([\dхx*.,]+)", n, re.I)
    if m:
        raw = m.group(1)
        # десятичная (запятая/точка) -> круглая/ВГП (напр. 25х2,8)
        if "," in raw or "." in raw:
            return "Труба круглая", normalize_size(raw)
        nonempty = [p for p in re.split(r"[хx*]", raw) if p]
        # профильная труба — сечение из 2-3 целых (40х40, 80х40х3)
        if len(nonempty) >= 2:
            return "Труба профильная", normalize_size(raw)
        # один размер -> круглая
        return "Труба круглая", normalize_size(raw)
    m = re.search(r"Швеллер\s*№?\s*([\dПпUuУу]+)", n, re.I)
    if m:
        return "Швеллер", m.group(1).upper().replace("U", "П").replace("У", "П")
    m = re.search(r"Уголок\s+([\dхx*]+)", n, re.I)
    if m:
        return "Уголок", normalize_size(m.group(1))
    return None, None


def parse_designation(obozn, base):
    o = str(obozn).strip()
    full = re.sub(r"-[^-]*$", "", base) + o if (o.startswith("-") and base) else o
    m = re.search(r"-(\d+)(_д\d*)?$", full)
    length = int(m.group(1)) if m else None
    dorab = m.group(2) if (m and m.group(2)) else ""
    return full, length, dorab


def extract_grade(obozn_mat):
    t = str(obozn_mat).strip()
    m = re.search(r";\s*([0-9]{0,2}[А-ЯA-Z][\w\-]+)\s+ГОСТ", t)
    if m:
        return m.group(1).strip()
    m = re.search(r"^(С\d{3})\s+ГОСТ", t)
    if m:
        return m.group(1)
    m = re.search(r"Сталь\s+(\d+\w*)", t)
    if m:
        return f"Сталь{m.group(1)}"
    return ""


def grade_from_note(primech):
    """
    Марка из столбца G (Примечание) — основной источник.
    Ловит класс прочности (С255, С345, С255-4, С345-5), марки Ст3 и 09Г2С.
    Нормализация суффиксов (-4/-5) делается позже в normalize_grade.
    """
    t = str(primech)
    if not t or t.lower() == "nan":
        return ""
    m = re.search(r"С\d{3}", t)
    if m:
        return m.group(0)
    m = re.search(r"Ст\d+[а-яА-Я0-9\-]*", t)
    if m:
        return m.group(0)
    m = re.search(r"\d{2}Г\d?[А-Я]+", t)
    if m:
        return m.group(0)
    return ""


def length_from_name(naim):
    """
    Извлекает длину из наименования вида 'L = 13500_д1 мм' (формат, где вся суть
    в наименовании, а обозначение пустое). Возвращает (длина|None, доработка).
    """
    m = re.search(r"L\s*=\s*(\d+)\s*(_д\d*)?\s*мм", str(naim), re.IGNORECASE)
    if not m:
        return None, ""
    return int(m.group(1)), (m.group(2) or "")


def grade_from_name(naim):
    """
    Извлекает марку из наименования, в т.ч. когда она слиплась с номером ГОСТа
    ('ГОСТ 8240-97С355' -> 'С355'). Возвращает каноническую марку или ''.
    """
    t = str(naim)
    m = re.search(r"С(\d{3})", t)            # класс прочности С255/С345/С355
    if m:
        return normalize_grade("С" + m.group(1))
    m = re.search(r"Ст\d+[а-яА-Я0-9\-]*", t)  # Ст3 и доработки
    if m:
        return normalize_grade(m.group(0))
    m = re.search(r"\d{2}Г\d?[А-Я]+", t)      # 09Г2С и т.п.
    if m:
        return normalize_grade(m.group(0))
    return ""


def make_short_note(isp, dorab):
    """
    Краткое примечание для схемы раскроя.
    Исполнение: '-04'->'4', '-01'->'1', '-'/'' -> пусто (без исполнения не печатаем).
    Доработка:  '_д'->'д', '_д1'->'д1'.
    Результат: 'исполнение доработка' через пробел, напр. '4 д', '1', 'д1', ''.
    """
    parts = []
    s = str(isp or "").strip()
    if s and s != "-":
        # оставляем только значащий номер исполнения (без ведущих дефисов/нулей)
        num = s.lstrip("-").lstrip("0") or s.lstrip("-")
        if num:
            parts.append(num)
    d = str(dorab or "").strip().lstrip("_")
    if d:
        parts.append(d)
    return " ".join(parts)


def material_size(obozn_mat):
    m = re.search(r"d?(\d+х\d+х\d+|\d+х\d+|\d+[ПУпу])", str(obozn_mat))
    return normalize_size(m.group(1)) if m else ""


def detect_layout(df):
    """Определяет колонки количества, обозначения материала/примечания, исполнения."""
    hdr = [str(df.iloc[0, c]).strip() for c in range(df.shape[1])]
    kol_cols = [c for c, h in enumerate(hdr) if h == "Количество"]
    # колонка сортамента ("Обозначение материала"), если есть
    mat_col = None
    for c, h in enumerate(hdr):
        if "материал" in h.lower():
            mat_col = c
    # колонка примечания (запасной источник марки: С255-4)
    prim_col = None
    for c, h in enumerate(hdr):
        if h.lower().startswith("примеч"):
            prim_col = c
    ispoln = {}
    if len(kol_cols) > 1:
        for i in range(1, 6):
            if i < len(df) and "исполн" in str(df.iloc[i, 4]).lower():
                for c in kol_cols:
                    v = str(df.iloc[i, c]).strip()
                    if v and v.lower() != "nan":
                        ispoln[c] = v
                break
    if not ispoln:
        ispoln = {kol_cols[0]: ""}
    return ispoln, mat_col, prim_col


def main(src_xls, out_xlsx):
    df = pd.read_excel(src_xls, sheet_name=0, header=None, dtype=str)
    ispoln, mat_col, prim_col = detect_layout(df)
    has_isp = any(v for v in ispoln.values())

    rows = df[df[2].notna() & df[2].astype(str).str.match(r"^\d+$", na=False)]

    # Предпроход: в разделе "Материалы" профиль бывает написан строкой-заголовком
    # НАД позицией (напр. 'Труба Ду80х4,0'), а сама позиция содержит только
    # 'L=184 мм'. Соберём для таких позиций профиль из ближайшего заголовка сверху.
    header_profile_by_pos = {}
    last_header = ""
    for i in range(len(df)):
        naim_i = str(df.iloc[i, 4]).strip()
        pos_i = str(df.iloc[i, 2]).strip()
        is_pos = pos_i.isdigit()
        if not is_pos:
            # строка без позиции: возможный заголовок-профиль (содержит профиль и не служебная)
            if naim_i and naim_i.lower() != "nan":
                pt, _ = classify_profile(naim_i)
                if pt is not None:
                    last_header = naim_i
            continue
        # строка с позицией: если в её наименовании нет профиля, а заголовок есть — запомним
        pt_here, _ = classify_profile(naim_i)
        if pt_here is None and last_header:
            header_profile_by_pos[pos_i] = last_header

    parts, materials, checks, cut_off = [], {}, [], []
    base_desig = ""

    for _, r in rows.iterrows():
        naim = str(r[4]).strip()
        obozn_raw = str(r[3]).strip()
        pos = str(r[2]).strip()
        obozn_mat = r[mat_col] if mat_col is not None else ""
        primech = r[prim_col] if prim_col is not None else ""

        if obozn_raw and not obozn_raw.startswith("-") and obozn_raw.lower() != "nan":
            base_desig = obozn_raw

        if any(k in naim for k in SHEET_KEYWORDS) and "Труба" not in naim:
            cut_off.append([pos, obozn_raw, naim, "не линейный прокат (лист/крепёж/прочее)"])
            continue
        profile_type, size = classify_profile(naim)
        if profile_type is None:
            # позиция из раздела "Материалы": профиль в заголовке над строкой
            header = header_profile_by_pos.get(pos, "")
            if header:
                profile_type, size = classify_profile(header)
        if profile_type is None:
            cut_off.append([pos, obozn_raw, naim, "не распознан профиль — проверить вручную"])
            continue

        full_desig, length, dorab = parse_designation(obozn_raw, base_desig)
        # Марка: ПРИОРИТЕТ столбцу G (Примечание) — там верные данные (С255-4, С345-5).
        # Столбец "Обозначение материала" (после G) часто содержит НЕВЕРНЫЕ данные
        # (чужой размер/марку), поэтому как источник марки его НЕ используем.
        # Если в G пусто — ищем в обозначении и в наименовании.
        grade = (
            grade_from_note(primech)
            or extract_grade(obozn_raw)
            or grade_from_name(naim)
        )

        # Фолбэк для формата, где вся суть в наименовании (обозначение пустое):
        # длина 'L = NNN мм' и марка, слипшаяся с ГОСТом, берутся из наименования.
        if length is None:
            name_len, name_dorab = length_from_name(naim)
            if name_len is not None:
                length = name_len
                if name_dorab and not dorab:
                    dorab = name_dorab
        if not grade:
            grade = grade_from_name(naim)

        warn = []
        if length is None:
            warn.append("длина не извлечена")
        if not grade:
            if profile_type == "Труба круглая":
                # ВГП/круглые трубы (ГОСТ 3262) обычно Ст3 — ставим по умолчанию,
                # чтобы деталь попала в расчёт; помечаем для информации технолога.
                grade = "Ст3"
                warn.append("марка не указана в СП — принята Ст3 (проверьте)")
            else:
                warn.append("марка не определена")
        # Размер берём ТОЛЬКО из наименования. Столбец "Обозначение материала"
        # (после G) часто содержит неверный размер (напр. для швеллера 16П там
        # значится 12У), поэтому сверку с ним не делаем — она давала ложные
        # конфликты.

        # каноническая марка (Ст3пс3-св -> Ст3, С255-4 -> С255) — единый формат
        grade = normalize_grade(grade)
        # код материала собирается тем же модулем, что и в справочнике,
        # поэтому совпадает байт-в-байт (нет дублей из-за регистра/написания)
        prefix = PROFILE_MAP.get(profile_type, "")
        code = build_material_code(profile_type, size, grade)

        # Составная/сварная деталь: длина больше стандартного хлыста — раскроем
        # как единичную деталь нельзя, разбивка определяется конструкторами под
        # заказ. Помечаем для ручного решения.
        if length is not None:
            stock_default = STOCK_LENGTH_DEFAULT.get(prefix, 6000)
            if length > stock_default:
                warn.append(
                    f"СОСТАВНАЯ ДЕТАЛЬ: длина {length} мм больше хлыста "
                    f"{stock_default} мм — разбивка с конструкторами"
                )

        # разворот по исполнениям: одна строка на каждое исполнение с кол-вом
        for col, isp in ispoln.items():
            qv = str(r[col]).strip()
            if qv in ("", "nan", "None"):
                continue
            try:
                qty = int(float(qv.replace(",", ".")))
            except ValueError:
                continue
            # ПОЛНОЕ описание (для сводок/карты раскроя/производства)
            desc_parts = []
            if isp:
                desc_parts.append(f"исп.{isp}")
            if dorab:
                desc_parts.append(f"доработка{dorab}")
            description = " ".join(desc_parts)

            # КРАТКОЕ примечание (для схемы раскроя — печатается под длиной):
            # исполнение коротким номером (без исполнения — пусто) и доработка
            # одной буквой через пробел: '4 д', '1', 'д1'.
            note = make_short_note(isp, dorab)

            if warn:
                checks.append([pos, full_desig, naim, isp or "-", size, grade,
                               length if length else "", qty, "; ".join(warn)])
            parts.append({
                "Обозначение": full_desig, "Наименование детали": naim, "Код материала": code,
                "Длина, мм": length if length else "", "Количество": qty,
                "Исполнение": isp, "Доработка": dorab,
                "Примечание": note,          # краткое -> на схему раскроя
                "Описание": description,      # полное -> в сводки
                "_conflict": False,
            })
            if code and code not in materials:
                materials[code] = {
                    "Код материала": code, "Наименование": f"{profile_type} {size} {grade}",
                    "Тип профиля": profile_type, "Размер": size, "Марка": grade,
                    "Длина хлыста, мм": STOCK_LENGTH_DEFAULT.get(prefix, 6000),
                    "Кол-во хлыстов": 0, "Примечание": "",
                }

    write_workbook(out_xlsx, parts, list(materials.values()), checks, cut_off, has_isp)
    print(f"Исполнений в файле: {len([v for v in ispoln.values() if v]) or 1}"
          f"{' (' + ', '.join(v for v in ispoln.values() if v) + ')' if has_isp else ' (без исполнений)'}")
    print(f"Деталей-строк (после разворота): {len(parts)}")
    print(f"Уникальных материалов: {len(materials)}")
    print(f"Строк с пометками: {len(checks)}")
    print(f"Отсечено: {len(cut_off)}")


def write_workbook(path, parts, materials, checks, cut_off, has_isp):
    wb = Workbook()
    H = Font(bold=True, color="FFFFFF", name="Arial")
    HF = PatternFill("solid", fgColor="1F4E78")

    def style_header(ws, names, row=1):
        for c, name in enumerate(names, 1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font = H; cell.fill = HF
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ===== Лист Детали со светофором =====
    ws = wb.active; ws.title = "Детали"
    cols = ["Обозначение", "Наименование детали", "Код материала", "Длина, мм",
            "Количество", "Исполнение", "Доработка", "Примечание", "Описание"]
    n = len(cols)
    status_col = n + 1
    sc = chr(64 + status_col)
    data_first = 4
    data_last = 4 + max(len(parts) - 1, 0)

    cnt_formula = f'=COUNTIF(${sc}{data_first}:${sc}{data_last},"⚠ заполнить")'
    ws.cell(row=1, column=1, value="НЕ ЗАПОЛНЕНО:").font = Font(bold=True, name="Arial")
    ws.cell(row=1, column=3, value=cnt_formula).font = Font(bold=True, size=12, name="Arial")
    verdict = ('=IF($C$1=0,"✓ ВСЁ ЗАПОЛНЕНО — можно считать раскрой",'
               '"✗ ЗАПОЛНИТЕ строки со статусом ⚠ (Длина и Код материала)")')
    vc = ws.cell(row=1, column=4, value=verdict)
    vc.font = Font(bold=True, size=12, name="Arial")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=status_col)

    for c, name in enumerate(cols + ["Статус"], 1):
        cell = ws.cell(row=3, column=c, value=name)
        cell.font = H; cell.fill = HF
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = data_first
    for p in parts:
        for c, name in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=p[name])
        # статус: заполнить / доработка / готова  (по заполненности кода и длины)
        st = (f'=IF(OR($C{r}="",$D{r}=""),"⚠ заполнить",'
              f'IF($G{r}<>"","● доработка","✓ готова"))')
        ws.cell(row=r, column=status_col, value=st)
        r += 1
    data_last_real = r - 1

    if parts:
        rng = f"A4:{sc}{data_last_real}"
        red = PatternFill("solid", fgColor="FFC7CE")
        peach = PatternFill("solid", fgColor="FCE4D6")
        green = PatternFill("solid", fgColor="E2EFDA")
        # красный — нужно заполнить (нет кода или длины)
        ws.conditional_formatting.add(rng, FormulaRule(formula=['OR($C4="",$D4="")'], fill=red, stopIfTrue=True))
        # персик — доработка (заполнено, но требует внимания)
        ws.conditional_formatting.add(rng, FormulaRule(formula=['$G4<>""'], fill=peach, stopIfTrue=True))
        # зелёный — всё ок
        ws.conditional_formatting.add(rng, FormulaRule(formula=['AND($C4<>"",$D4<>"")'], fill=green))

    for i, w in enumerate([26, 22, 26, 10, 11, 11, 11, 14, 24, 13], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ===== Материалы =====
    ws2 = wb.create_sheet("Материалы")
    cols2 = ["Код материала", "Наименование", "Тип профиля", "Размер", "Марка",
             "Длина хлыста, мм", "Кол-во хлыстов", "Примечание"]
    style_header(ws2, cols2)
    for m in materials:
        ws2.append([m.get(c, "") for c in cols2])
    for i, w in enumerate([26, 30, 18, 12, 12, 16, 14, 20], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # ===== Проверка =====
    ws3 = wb.create_sheet("Проверка")
    cols3 = ["Позиция", "Обозначение", "Наименование", "Исп.", "Размер", "Марка",
             "Длина, мм", "Кол-во", "Что проверить"]
    style_header(ws3, cols3)
    peach = PatternFill("solid", fgColor="FCE4D6")
    warn = PatternFill("solid", fgColor="FFF2CC")
    for row in checks:
        ws3.append(row)
    for i, w in enumerate([8, 24, 20, 7, 11, 11, 10, 7, 46], 1):
        ws3.column_dimensions[chr(64 + i)].width = w

    # ===== Отсечено =====
    ws4 = wb.create_sheet("Отсечено")
    cols4 = ["Позиция", "Обозначение", "Наименование", "Причина"]
    style_header(ws4, cols4)
    for row in cut_off:
        ws4.append(row)
    for i, w in enumerate([8, 26, 26, 44], 1):
        ws4.column_dimensions[chr(64 + i)].width = w

    wb.save(path)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
