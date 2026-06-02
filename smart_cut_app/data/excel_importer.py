from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from core.models import Material, Part


class ExcelImportError(Exception):
    pass


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_header_map(sheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for column_index, cell in enumerate(sheet[1], start=1):
        header = _normalize_header(cell.value)
        if header:
            headers[header] = column_index
    return headers


def _get_cell_value(sheet, row_index: int, column_index: int) -> str:
    value = sheet.cell(row=row_index, column=column_index).value
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: str) -> int:
    if not value:
        return 0
    try:
        return int(float(value.replace(",", ".")))
    except ValueError:
        return 0


def _to_float(value: str) -> float:
    if not value:
        return 0.0
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return 0.0


def import_materials_from_excel(file_path: str) -> List[Material]:
    path = Path(file_path)
    if not path.exists():
        raise ExcelImportError(f"Файл не найден: {file_path}")

    workbook = load_workbook(filename=file_path, data_only=True)

    if "Материалы" not in workbook.sheetnames:
        raise ExcelImportError("В файле отсутствует лист 'Материалы'.")

    sheet = workbook["Материалы"]
    header_map = _get_header_map(sheet)

    required_headers = [
        "Код материала",
        "Наименование",
        "Тип профиля",
        "Размер",
        "Длина хлыста, мм",
    ]

    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ExcelImportError(
            "На листе 'Материалы' отсутствуют обязательные колонки: "
            + ", ".join(missing_headers)
        )

    materials: List[Material] = []

    for row_index in range(2, sheet.max_row + 1):
        code = _get_cell_value(sheet, row_index, header_map["Код материала"])
        name = _get_cell_value(sheet, row_index, header_map["Наименование"])
        profile_type = _get_cell_value(sheet, row_index, header_map["Тип профиля"])
        size = _get_cell_value(sheet, row_index, header_map["Размер"])
        grade = _get_cell_value(sheet, row_index, header_map.get("Марка", 0)) if "Марка" in header_map else ""
        stock_length_mm = _to_int(
            _get_cell_value(sheet, row_index, header_map["Длина хлыста, мм"])
        )
        mass_per_meter = _to_float(
            _get_cell_value(sheet, row_index, header_map.get("Масса 1 м, кг", 0))
        ) if "Масса 1 м, кг" in header_map else 0.0
        price_per_meter = _to_float(
            _get_cell_value(sheet, row_index, header_map.get("Цена за 1 м", 0))
        ) if "Цена за 1 м" in header_map else 0.0
        available_count = _to_int(
            _get_cell_value(sheet, row_index, header_map.get("Кол-во хлыстов", 0))
        ) if "Кол-во хлыстов" in header_map else 0
        note = _get_cell_value(sheet, row_index, header_map.get("Примечание", 0)) if "Примечание" in header_map else ""

        if not any([code, name, profile_type, size, str(stock_length_mm), grade, note]):
            continue

        materials.append(
            Material(
                id=f"MAT-{len(materials) + 1:03d}",
                code=code,
                name=name,
                profile_type=profile_type,
                size=size,
                grade=grade,
                stock_length_mm=stock_length_mm,
                mass_per_meter=mass_per_meter,
                price_per_meter=price_per_meter,
                available_count=available_count,
                note=note,
            )
        )

    return materials


def import_parts_from_excel(file_path: str) -> List[Part]:
    path = Path(file_path)
    if not path.exists():
        raise ExcelImportError(f"Файл не найден: {file_path}")

    workbook = load_workbook(filename=file_path, data_only=True)

    if "Детали" not in workbook.sheetnames:
        raise ExcelImportError("В файле отсутствует лист 'Детали'.")

    sheet = workbook["Детали"]
    header_map = _get_header_map(sheet)

    required_headers = [
        "Наименование детали",
        "Код материала",
        "Длина, мм",
        "Количество",
    ]

    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ExcelImportError(
            "На листе 'Детали' отсутствуют обязательные колонки: "
            + ", ".join(missing_headers)
        )

    parts: List[Part] = []

    for row_index in range(2, sheet.max_row + 1):
        designation = _get_cell_value(sheet, row_index, header_map.get("Обозначение", 0)) if "Обозначение" in header_map else ""
        name = _get_cell_value(sheet, row_index, header_map["Наименование детали"])
        material_code = _get_cell_value(sheet, row_index, header_map["Код материала"])
        length_mm = _to_int(_get_cell_value(sheet, row_index, header_map["Длина, мм"]))
        quantity = _to_int(_get_cell_value(sheet, row_index, header_map["Количество"]))
        assembly = _get_cell_value(sheet, row_index, header_map.get("Узел", 0)) if "Узел" in header_map else ""
        note = _get_cell_value(sheet, row_index, header_map.get("Примечание", 0)) if "Примечание" in header_map else ""

        if not any([designation, name, material_code, str(length_mm), str(quantity), assembly, note]):
            continue

        parts.append(
            Part(
                id=f"PART-{len(parts) + 1:03d}",
                designation=designation,
                name=name,
                material_code=material_code,
                length_mm=length_mm,
                quantity=quantity,
                assembly=assembly,
                note=note,
            )
        )

    return parts
