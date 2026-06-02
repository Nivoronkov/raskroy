from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import CalculationResult


class ExcelExportError(Exception):
    pass


THIN_BORDER = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)

TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
DETAIL_FILLS = [
    PatternFill("solid", fgColor="FFF2CC"),
    PatternFill("solid", fgColor="DDEBF7"),
    PatternFill("solid", fgColor="E2F0D9"),
    PatternFill("solid", fgColor="FCE4D6"),
]
LEFTOVER_FILL = PatternFill("solid", fgColor="E7E6E6")


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _display_group_key(designation: str, name: str) -> str:
    designation = designation.strip() if designation else ""
    name = name.strip() if name else ""

    if not designation:
        designation = "Без обозначения"
    if not name:
        name = "Без наименования"

    return f"{designation} — {name}"


def _get_pattern_group_title(pattern) -> str:
    group_titles = set()

    for part in pattern.parts:
        group_titles.add(
            _display_group_key(
                getattr(part, "designation", ""),
                getattr(part, "name", ""),
            )
        )

    if not group_titles:
        return "Без обозначения — Без наименования"

    if len(group_titles) == 1:
        return next(iter(group_titles))

    return "Смешанная группа"


def _set_header_row(sheet, headers: list[str], row_num: int) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_num, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _build_visual_cell_text(part) -> str:
    lines = [str(part.base_length_mm)]

    assembly = getattr(part, "assembly", "")
    note = getattr(part, "note", "")

    if assembly:
        assembly = str(assembly).strip()
    else:
        assembly = ""

    if note:
        note = str(note).strip()
    else:
        note = ""

    if assembly:
        lines.append(assembly)

    if note:
        lines.append(note)

    return "\n".join(lines[:3])


def _fill_visual_pattern(sheet, start_row: int, pattern, pattern_index: int) -> int:
    """
    Рисует схему одного хлыста.
    Возвращает номер следующей свободной строки.
    """
    total_visual_cols = 60
    first_visual_col = 2
    last_visual_col = first_visual_col + total_visual_cols - 1

    source_text = "Остаток" if pattern.source_type == "leftover" else "Новый хлыст"
    title = (
        f"{source_text} #{pattern_index} | "
        f"{pattern.material_code} | {pattern.material_name} | "
        f"Хлыст: {pattern.stock_length_mm} мм | "
        f"Остаток: {pattern.leftover_length_mm} мм"
    )

    # Заголовок
    sheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=last_visual_col,
    )
    title_cell = sheet.cell(start_row, 1, title)
    title_cell.font = Font(bold=True)
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[start_row].height = 22

    # Полоса схемы
    for col in range(1, last_visual_col + 1):
        sheet.cell(start_row + 1, col).border = THIN_BORDER

    sheet.cell(start_row + 1, 1, "Схема")
    sheet.cell(start_row + 1, 1).font = Font(bold=True)
    sheet.cell(start_row + 1, 1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(start_row + 1, 1).border = THIN_BORDER
    sheet.row_dimensions[start_row + 1].height = 48

    used_cols = 0
    current_col = first_visual_col
    stock_length = max(pattern.stock_length_mm, 1)

    for idx, part in enumerate(pattern.parts):
        length = int(part.base_length_mm)
        width_cols = max(1, round((length / stock_length) * total_visual_cols))

        remaining_cols = total_visual_cols - used_cols
        width_cols = min(width_cols, max(1, remaining_cols))
        end_col = current_col + width_cols - 1

        if current_col > last_visual_col:
            break

        if end_col > last_visual_col:
            end_col = last_visual_col

        sheet.merge_cells(
            start_row=start_row + 1,
            start_column=current_col,
            end_row=start_row + 1,
            end_column=end_col,
        )

        cell_text = _build_visual_cell_text(part)
        cell = sheet.cell(start_row + 1, current_col, cell_text)
        cell.fill = DETAIL_FILLS[idx % len(DETAIL_FILLS)]
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER

        for c in range(current_col, end_col + 1):
            sheet.cell(start_row + 1, c).border = THIN_BORDER
            sheet.cell(start_row + 1, c).fill = DETAIL_FILLS[idx % len(DETAIL_FILLS)]

        segment_width = end_col - current_col + 1
        used_cols += segment_width
        current_col = end_col + 1

    # Остаток
    leftover_cols = total_visual_cols - used_cols
    if leftover_cols > 0 and current_col <= last_visual_col:
        end_col = last_visual_col

        sheet.merge_cells(
            start_row=start_row + 1,
            start_column=current_col,
            end_row=start_row + 1,
            end_column=end_col,
        )
        leftover_text = f"Ост.\n{pattern.leftover_length_mm}"
        cell = sheet.cell(start_row + 1, current_col, leftover_text)
        cell.fill = LEFTOVER_FILL
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER

        for c in range(current_col, end_col + 1):
            sheet.cell(start_row + 1, c).border = THIN_BORDER
            sheet.cell(start_row + 1, c).fill = LEFTOVER_FILL

    # Подпись снизу
    part_lengths = [str(part.base_length_mm) for part in pattern.parts]
    summary_text = (
        f"Детали: {' + '.join(part_lengths)} | "
        f"Занято: {pattern.used_length_mm} мм | "
        f"Остаток: {pattern.leftover_length_mm} мм | "
        f"Резов: {pattern.cuts_count}"
    )
    sheet.merge_cells(
        start_row=start_row + 2,
        start_column=1,
        end_row=start_row + 2,
        end_column=last_visual_col,
    )
    bottom_cell = sheet.cell(start_row + 2, 1, summary_text)
    bottom_cell.alignment = Alignment(vertical="center")
    bottom_cell.font = Font(italic=True)
    sheet.row_dimensions[start_row + 2].height = 22

    return start_row + 4

def export_result_to_excel(result: CalculationResult, file_path: str) -> None:
    if not result.success:
        raise ExcelExportError("Нельзя экспортировать результат с ошибками расчета.")

    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise ExcelExportError("Файл экспорта должен иметь расширение .xlsx")

    workbook = Workbook()

    # Лист 1: Сводка
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"

    summary_headers = [
        "Код материала",
        "Материал",
        "Длина хлыста, мм",
        "Кол-во деталей",
        "Кол-во хлыстов",
        "Длина деталей, мм",
        "Потери на рез, мм",
        "Отход, мм",
        "Полезный остаток, мм",
        "Использование, %",
    ]
    summary_sheet.append(summary_headers)
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    for row in result.summary_rows:
        summary_sheet.append([
            row.material_code,
            row.material_name,
            row.stock_length_mm,
            row.total_parts_count,
            row.used_bars_count,
            row.total_parts_length_mm,
            row.total_cut_loss_mm,
            row.total_waste_mm,
            row.total_useful_leftover_mm,
            row.utilization_percent,
        ])

    _autosize_columns(summary_sheet)

    # Лист 2: Карты раскроя
    patterns_sheet = workbook.create_sheet("Карты раскроя")

    pattern_headers = [
        "№ хлыста",
        "Источник",
        "Код материала",
        "Материал",
        "Длина хлыста, мм",
        "Состав раскроя",
        "Резов",
        "Занято, мм",
        "Остаток, мм",
        "Тип остатка",
    ]
    patterns_sheet.append(pattern_headers)
    for cell in patterns_sheet[1]:
        cell.font = Font(bold=True)

    for index, pattern in enumerate(result.patterns, start=1):
        source_text = "Остаток" if pattern.source_type == "leftover" else "Новый хлыст"
        patterns_sheet.append([
            index,
            source_text,
            pattern.material_code,
            pattern.material_name,
            pattern.stock_length_mm,
            pattern.pattern_as_text(),
            pattern.cuts_count,
            pattern.used_length_mm,
            pattern.leftover_length_mm,
            pattern.leftover_type,
        ])

    _autosize_columns(patterns_sheet)

    # Лист 3: Производство
    production_sheet = workbook.create_sheet("Производство")
    production_sheet["A1"] = 'Справочно: для выдачи задания в цех использовать лист "Карты раскроя".'
    production_sheet["A1"].font = Font(bold=True)
    production_sheet.append([])

    production_headers = [
        "Источник",
        "Код материала",
        "Материал",
        "Длина хлыста, мм",
        "Кол-во одинаковых хлыстов",
        "Схема раскроя",
        "Резов",
        "Занято, мм",
        "Остаток, мм",
        "Тип остатка",
    ]
    production_sheet.append(production_headers)
    for cell in production_sheet[3]:
        cell.font = Font(bold=True)

    grouped = {}
    for pattern in result.patterns:
        group_title = _get_pattern_group_title(pattern)
        grouped.setdefault(group_title, []).append(pattern)

    for group_title, patterns in grouped.items():
        production_sheet.append([group_title])
        current_row = production_sheet.max_row
        production_sheet.cell(current_row, 1).font = Font(bold=True)
        production_sheet.cell(current_row, 1).fill = TITLE_FILL

        for pattern in patterns:
            source_text = "Остаток" if pattern.source_type == "leftover" else "Новый хлыст"
            production_sheet.append([
                source_text,
                pattern.material_code,
                pattern.material_name,
                pattern.stock_length_mm,
                1,
                pattern.pattern_as_text(),
                pattern.cuts_count,
                pattern.used_length_mm,
                pattern.leftover_length_mm,
                pattern.leftover_type,
            ])

    _autosize_columns(production_sheet)

    # Лист 4: Движение остатков
    movement_sheet = workbook.create_sheet("Движение остатков")

    movement_headers = [
        "Операция",
        "ID остатка",
        "Код материала",
        "Материал",
        "Длина, мм",
        "Примечание",
    ]
    movement_sheet.append(movement_headers)
    for cell in movement_sheet[1]:
        cell.font = Font(bold=True)

    for row in result.leftover_movements:
        movement_sheet.append([
            row.operation_type,
            row.leftover_id,
            row.material_code,
            row.material_name,
            row.length_mm,
            row.note,
        ])

    _autosize_columns(movement_sheet)

    # Лист 5: Новые остатки
    new_leftovers_sheet = workbook.create_sheet("Новые остатки")

    new_leftovers_headers = [
        "ID остатка",
        "Код материала",
        "Материал",
        "Длина остатка, мм",
        "Исходная длина заготовки, мм",
        "Источник",
        "Примечание",
    ]
    new_leftovers_sheet.append(new_leftovers_headers)
    for cell in new_leftovers_sheet[1]:
        cell.font = Font(bold=True)

    for row in result.leftovers:
        new_leftovers_sheet.append([
            row.id,
            row.material_code,
            row.material_name,
            row.length_mm,
            row.stock_length_mm,
            row.source_pattern_id,
            row.note,
        ])

    _autosize_columns(new_leftovers_sheet)

    # Лист 6: Схема раскроя
    visual_sheet = workbook.create_sheet("Схема раскроя")

    # Ширины колонок под визуализацию
    visual_sheet.column_dimensions["A"].width = 16
    for col_index in range(2, 62):
        visual_sheet.column_dimensions[get_column_letter(col_index)].width = 3

    current_row = 1

    grouped_visual = {}
    for pattern in result.patterns:
        group_title = _get_pattern_group_title(pattern)
        grouped_visual.setdefault(group_title, []).append(pattern)

    pattern_counter = 1
    for group_title, patterns in grouped_visual.items():
        visual_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=61,
        )
        group_cell = visual_sheet.cell(current_row, 1, group_title)
        group_cell.font = Font(bold=True, size=12)
        group_cell.fill = TITLE_FILL
        group_cell.alignment = Alignment(vertical="center")
        current_row += 2

        for pattern in patterns:
            current_row = _fill_visual_pattern(
                visual_sheet,
                current_row,
                pattern,
                pattern_counter,
            )
            pattern_counter += 1

        current_row += 1

    workbook.save(file_path)