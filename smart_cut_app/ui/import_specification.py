"""
Импорт спецификации Компас (.xls) прямо в окно калькулятора.

Переиспользует конвертер sp_to_cutlist (тот же, что в командном raskroy.py):
спецификация -> временный файл "Вход" -> объекты Part для таблицы деталей.

Возвращает (parts, missing_codes, info_text):
  parts          — список Part для parts_tab.set_parts()
  missing_codes  — коды материалов, которых НЕТ в справочнике склада
  info_text      — текст-сводка для показа пользователю
"""
import os
import sys
import tempfile

from openpyxl import load_workbook
from core.models import Part
from data.materials_catalog_repository import load_materials_catalog

# Ищем sp_to_cutlist.py в нескольких местах и добавляем найденную папку в sys.path.
# Это надёжнее, чем один фиксированный путь: работает независимо от того,
# из какой папки запущен python main.py.
def _find_converter_dir():
    _ui = os.path.dirname(os.path.abspath(__file__))          # ...\smart_cut_app\ui
    _app = os.path.dirname(_ui)                                # ...\smart_cut_app
    _parent = os.path.dirname(_app)                            # ...\raskroy
    candidates = [
        _parent,                       # рядом с папкой smart_cut_app  (C:\raskroy)
        _app,                          # внутри smart_cut_app
        os.getcwd(),                   # текущая рабочая папка
        _ui,                           # на всякий случай — папка ui
    ]
    # в собранном .exe — папка рядом с исполняемым файлом
    if getattr(sys, "frozen", False):
        candidates.insert(0, os.path.dirname(sys.executable))
        # PyInstaller распаковывает данные во временную папку _MEIPASS
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidates.insert(0, meipass)
    for d in candidates:
        if os.path.exists(os.path.join(d, "sp_to_cutlist.py")):
            return d
    return None

_CONV_DIR = _find_converter_dir()
if _CONV_DIR and _CONV_DIR not in sys.path:
    sys.path.insert(0, _CONV_DIR)

HEADER_ROW = 3   # заголовки листа "Детали" в строке 3 (выше — вердикт-светофор)
DATA_ROW = 4


class SpecImportError(Exception):
    pass


def import_specification(xls_path: str):
    """Главная функция. Конвертирует спецификацию и возвращает детали."""
    try:
        from sp_to_cutlist import main as convert_sp
    except ImportError as exc:
        searched = _CONV_DIR or "(не найдено ни в одной из папок)"
        raise SpecImportError(
            "Не найден файл конвертера sp_to_cutlist.py.\n"
            f"Искал рядом с smart_cut_app, внутри неё и в текущей папке.\n"
            f"Найденная папка: {searched}\n"
            "Положи sp_to_cutlist.py рядом с папкой smart_cut_app "
            "(в корневую папку программы)."
        ) from exc

    if not os.path.exists(xls_path):
        raise SpecImportError(f"Файл не найден: {xls_path}")

    # конвертируем во временный xlsx
    tmp = tempfile.NamedTemporaryFile(suffix="_вход.xlsx", delete=False)
    tmp.close()
    try:
        convert_sp(xls_path, tmp.name)
        parts = _read_parts(tmp.name)
        spec_materials = _read_materials(tmp.name)   # карточки из листа "Материалы"
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    if not parts:
        raise SpecImportError(
            "В спецификации не найдено деталей линейного проката "
            "(труб, швеллеров, уголков)."
        )

    missing = _check_catalog(parts)
    # карточки только недостающих материалов — для быстрого добавления на склад
    missing_cards = [m for m in spec_materials if m["material_code"] in set(missing)]
    info = _build_info(parts, missing)
    return parts, missing, info, missing_cards


def _read_materials(vhod_path: str):
    """
    Читает лист 'Материалы' временного файла -> список карточек:
    {material_code, profile_type, size, grade, stock_length_mm}.
    Нужно для предзаполнения формы быстрого добавления на склад.
    """
    wb = load_workbook(vhod_path, data_only=True)
    if "Материалы" not in wb.sheetnames:
        return []
    ws = wb["Материалы"]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c

    def cell(r, name):
        c = headers.get(name)
        return ws.cell(row=r, column=c).value if c else None

    cards = []
    for r in range(2, ws.max_row + 1):
        code = cell(r, "Код материала")
        if not code:
            continue
        cards.append({
            "material_code": str(code).strip(),
            "profile_type": str(cell(r, "Тип профиля") or "").strip(),
            "size": str(cell(r, "Размер") or "").strip(),
            "grade": str(cell(r, "Марка") or "").strip(),
            "stock_length_mm": _to_int(cell(r, "Длина хлыста, мм")) or 6000,
        })
    return cards



def _read_parts(vhod_path: str):
    """Читает лист 'Детали' (с учётом сдвига заголовков) -> список Part."""
    wb = load_workbook(vhod_path, data_only=True)
    ws = wb["Детали"]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            headers[str(v).strip()] = c

    def cell(r, name):
        c = headers.get(name)
        return ws.cell(row=r, column=c).value if c else None

    parts = []
    idx = 0
    for r in range(DATA_ROW, ws.max_row + 1):
        obozn = cell(r, "Обозначение")
        naim = cell(r, "Наименование детали")
        if not obozn and not naim:
            continue
        idx += 1
        length = cell(r, "Длина, мм")
        qty = cell(r, "Количество")
        parts.append(Part(
            id=f"PART-{idx:03d}",
            designation=str(obozn or ""),
            name=str(naim or ""),
            material_code=str(cell(r, "Код материала") or ""),
            length_mm=_to_int(length),
            quantity=_to_int(qty),
            note=str(cell(r, "Примечание") or ""),   # исполнение/доработка -> на карту
        ))
    return parts


def _check_catalog(parts):
    """Какие material_code из деталей отсутствуют в справочнике склада."""
    catalog = {item.material_code for item in load_materials_catalog() if item.is_active}
    missing = []
    for p in parts:
        code = (p.material_code or "").strip()
        if code and code not in catalog and code not in missing:
            missing.append(code)
    return missing


def _build_info(parts, missing):
    lines = [f"Загружено деталей: {len(parts)}"]
    no_code = [p for p in parts if not p.material_code.strip()]
    no_len = [p for p in parts if p.length_mm <= 0]
    if no_code:
        lines.append(f"⚠ Без кода материала: {len(no_code)} (заполните вручную)")
    if no_len:
        lines.append(f"⚠ Без длины: {len(no_len)} (заполните вручную)")
    if missing:
        lines.append("")
        lines.append(f"Нет в справочнике склада ({len(missing)}):")
        for code in missing:
            lines.append(f"   • {code}")
        lines.append("Добавьте их на вкладке «Справочник материалов» "
                     "(указав длины хлыстов), иначе они не попадут в расчёт.")
    elif not no_code and not no_len:
        lines.append("✓ Все материалы есть в справочнике. Можно рассчитывать.")
    return "\n".join(lines)


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", ".")))
    except ValueError:
        return 0
