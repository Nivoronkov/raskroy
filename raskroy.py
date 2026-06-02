"""
РАСКРОЙ ИЗ СПЕЦИФИКАЦИИ — скрипт "всё в одном".

Процесс:
  1. Читает выгрузку спецификации Компас (.xls), конвертирует (sp_to_cutlist)
  2. ПРОВЕРКА ГОТОВНОСТИ: если есть незаполненные строки (нет длины или кода
     материала) — НЕ считает, а сообщает какие строки заполнить.
  3. Сверяется со справочником склада (materials_catalog.json)
  4. Считает раскрой, сохраняет карту в Excel.

Запуск:
    python raskroy.py "спецификация.xls"

Если файл "_вход.xlsx" уже существует и ты его поправил вручную —
скрипт возьмёт твои правки (повторный запуск читает существующий _вход).
Чтобы пересоздать _вход с нуля из спецификации, удали файл _вход.xlsx.
"""
import os
import sys
import json

HERE = os.path.dirname(os.path.abspath(__file__))
CALC_DIR = os.path.join(HERE, "smart_cut_app")
sys.path.insert(0, CALC_DIR)
sys.path.insert(0, HERE)

from openpyxl import load_workbook
from sp_to_cutlist import main as convert_sp
from data.excel_importer import import_materials_from_excel
from core.cutting_engine import calculate_cutting
from core.models import CalculationSettings, Part, Material
from data.excel_exporter import export_result_to_excel

CATALOG_FILE = os.path.join(CALC_DIR, "materials_catalog.json")
HEADER_ROW = 3      # заголовки листа "Детали" в строке 3 (выше — вердикт)
DATA_ROW = 4        # данные с 4-й


def load_catalog_codes():
    if not os.path.exists(CATALOG_FILE):
        return {}
    with open(CATALOG_FILE, encoding="utf-8") as f:
        return {it["material_code"]: it for it in json.load(f)}


def read_parts_with_check(path):
    """
    Читает лист 'Детали' (с учётом сдвига) и проверяет готовность.
    Возвращает (parts, not_ready) где not_ready — список (строка, обозначение, чего нет).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["Детали"]
    # карта заголовков из строки HEADER_ROW
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            headers[str(v).strip()] = c

    def cell(r, name):
        c = headers.get(name)
        return ws.cell(row=r, column=c).value if c else None

    parts, not_ready = [], []
    idx = 0
    for r in range(DATA_ROW, ws.max_row + 1):
        obozn = cell(r, "Обозначение")
        naim = cell(r, "Наименование детали")
        if not obozn and not naim:
            continue
        code = cell(r, "Код материала")
        length = cell(r, "Длина, мм")
        qty = cell(r, "Количество")

        missing = []
        if not code:
            missing.append("Код материала")
        if not length:
            missing.append("Длина")
        if missing:
            not_ready.append((r, obozn or naim, ", ".join(missing)))
            continue

        idx += 1
        parts.append(Part(
            id=f"PART-{idx:03d}",
            designation=str(obozn or ""),
            name=str(naim or ""),
            material_code=str(code),
            length_mm=int(float(str(length).replace(",", "."))),
            quantity=int(float(str(qty).replace(",", "."))) if qty else 0,
            note=str(cell(r, "Примечание") or ""),
        ))
    return parts, not_ready


def main():
    if len(sys.argv) < 2:
        print('Использование: python raskroy.py "спецификация.xls"')
        sys.exit(1)

    src = sys.argv[1]
    if not os.path.exists(src):
        print(f"ОШИБКА: файл не найден: {src}")
        sys.exit(1)

    base = os.path.splitext(src)[0]
    f_in = base + "_вход.xlsx"
    f_out = base + "_раскрой.xlsx"

    print("=" * 70)
    print("ШАГ 1. Чтение спецификации")
    print("=" * 70)
    if os.path.exists(f_in):
        print(f"  Файл {os.path.basename(f_in)} уже существует — беру твои правки.")
        print("  (чтобы пересоздать из спецификации — удали этот файл)")
    else:
        convert_sp(src, f_in)

    print()
    print("=" * 70)
    print("ШАГ 2. Проверка готовности (заполнены ли длина и код материала)")
    print("=" * 70)
    parts, not_ready = read_parts_with_check(f_in)

    if not_ready:
        print(f"  ✗ НЕ ГОТОВО: {len(not_ready)} строк(и) без обязательных полей.")
        print(f"  Открой {os.path.basename(f_in)}, лист 'Детали', заполни красные строки:")
        for r, obozn, miss in not_ready:
            print(f"     строка {r}: {obozn}  —  нет: {miss}")
        print()
        print("  Заполни и запусти снова. Расчёт не выполнен.")
        sys.exit(2)

    print(f"  ✓ Готово: все {len(parts)} строк(и) заполнены. Можно считать.")

    print()
    print("=" * 70)
    print("ШАГ 3. Сверка материалов со справочником склада")
    print("=" * 70)
    catalog = load_catalog_codes()
    materials = import_materials_from_excel(f_in)
    missing = []
    for m in materials:
        if m.code in catalog:
            cat = catalog[m.code]
            m.stock_length_mm = cat.get("stock_length_mm", 6000) or 6000
            m.available_count = cat.get("available_stock_bars", 0)
            print(f"  [ЕСТЬ]  {m.code}  (склад: {m.available_count} хл. по {m.stock_length_mm} мм)")
        else:
            missing.append(m)
            if m.stock_length_mm == 0:
                m.stock_length_mm = 6000
            m.available_count = 999
            print(f"  [НЕТ В СПРАВОЧНИКЕ]  {m.code}  ->  добавить на склад")
    if missing:
        print(f"\n  ВНИМАНИЕ: {len(missing)} материал(ов) нет в справочнике "
              f"(взяты как хлыст 6000 мм — заглушка).")

    print()
    print("=" * 70)
    print("ШАГ 4. Расчёт раскроя")
    print("=" * 70)
    settings = CalculationSettings(
        cut_width_mm=2, trim_allowance_mm=0, min_useful_leftover_mm=300,
        optimization_mode="min_waste", use_leftovers=False,
    )
    res = calculate_cutting(materials, parts, settings)
    if not res.success:
        print("  ОШИБКИ расчёта:")
        for e in res.errors:
            print("   -", e)
        sys.exit(1)

    for row in res.summary_rows:
        print(f"  {row.material_code:26} хлыстов={row.used_bars_count:2}  "
              f"деталей={row.total_parts_count:3}  исп.={row.utilization_percent:5.1f}%  "
              f"отход={row.total_waste_mm} мм")

    export_result_to_excel(res, f_out)
    print()
    print("=" * 70)
    print("ГОТОВО")
    print("=" * 70)
    print(f"  Проверка обработки : {f_in}")
    print(f"  Карта раскроя      : {f_out}")


if __name__ == "__main__":
    main()
